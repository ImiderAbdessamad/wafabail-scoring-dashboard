"""Dump des feuilles Excel de référence (valeurs + formules) — outil d'analyse."""
from __future__ import annotations

import shutil
import sys
import tempfile
from pathlib import Path

import openpyxl

SRC = Path(
    r"c:\Users\Dell\OneDrive\Documents\ocr-cin\documentation-wafabail\use-case-Scoring"
    r"\fichiers Excel - Exemple de Scoring"
)


def dump(path: Path) -> None:
    for formulas in (True, False):
        wb = openpyxl.load_workbook(path, data_only=not formulas)
        tag = "FORMULES" if formulas else "VALEURS"
        for ws in wb.worksheets:
            print(f"\n===== {path.name} :: {ws.title} :: {tag} ({ws.max_row}x{ws.max_column}) =====")
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 90), max_col=min(ws.max_column, 8)):
                cells = []
                for c in row:
                    if c.value is None:
                        continue
                    cells.append(f"{c.coordinate}={c.value!r}")
                if cells:
                    print(" | ".join(cells))
        wb.close()


if __name__ == "__main__":
    for arg in sys.argv[1:]:
        dump(Path(arg))
